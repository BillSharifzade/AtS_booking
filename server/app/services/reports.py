"""Excel (.xlsx) export of booking history (Module C / Integration 4.2).

Produces a styled, detailed workbook: one row per booking with the full who/when/
where/what, plus a second sheet with the status-change history. Built in memory and
returned as bytes so it can be streamed by the API or sent as a Telegram document.
"""
from __future__ import annotations

from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import Booking, BookingStatus, Room

# Russian labels for the English enum (mirrors the frontend STATUS_LABELS).
RU_OUTCOME: dict[str, str] = {
    "held": "Состоялось",
    "partial": "Частично",
    "cancelled": "Отменено заказчиком",
}

RU_STATUS: dict[BookingStatus, str] = {
    BookingStatus.new: "Новая",
    BookingStatus.processing: "На обработке",
    BookingStatus.approved: "Подтверждена",
    BookingStatus.rejected: "Отклонена",
    BookingStatus.completed: "Завершена",
    BookingStatus.archived: "В архиве",
}

# Brand palette: deep blue header, white text.
_HEADER_FILL = PatternFill("solid", fgColor="1B2A4A")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_URGENT_FILL = PatternFill("solid", fgColor="FCE8E6")
_THIN = Side(style="thin", color="E0E4EA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP_TOP = Alignment(vertical="top", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="top")


def _as_utc(dt: datetime | None) -> datetime | None:
    if dt is None:
        return None
    return dt.replace(tzinfo=timezone.utc) if dt.tzinfo is None else dt


def _fmt_dt(dt: datetime | None) -> str:
    dt = _as_utc(dt)
    return dt.strftime("%d.%m.%Y %H:%M") if dt else "—"


# (header, width, alignment) — order defines the columns of the main sheet.
_COLUMNS: list[tuple[str, int, Alignment]] = [
    ("№", 6, _CENTER),
    ("Статус", 14, _WRAP_TOP),
    ("Срочная", 9, _CENTER),
    ("Зона", 14, _WRAP_TOP),
    ("Помещение", 20, _WRAP_TOP),
    ("Дата", 12, _CENTER),
    ("Начало", 8, _CENTER),
    ("Окончание", 10, _CENTER),
    ("Длит., ч", 9, _CENTER),
    ("Тип мероприятия", 20, _WRAP_TOP),
    ("Название", 28, _WRAP_TOP),
    ("Описание", 38, _WRAP_TOP),
    ("Участников", 11, _CENTER),
    ("Кофе-брейк", 18, _WRAP_TOP),
    ("Компания", 24, _WRAP_TOP),
    ("Контактное лицо", 22, _WRAP_TOP),
    ("Телефон", 18, _WRAP_TOP),
    ("Telegram ID", 14, _CENTER),
    ("Username", 18, _WRAP_TOP),
    ("Причина отклонения", 30, _WRAP_TOP),
    ("Итог мероприятия", 28, _WRAP_TOP),
    ("Оценка", 8, _CENTER),
    ("Отзыв", 34, _WRAP_TOP),
    ("Создана", 17, _CENTER),
    ("Обновлена", 17, _CENTER),
]


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


_COFFEE_TYPE_RU = {"standard": "стандартный", "other": "другое"}


def _coffee_text(b: Booking) -> str:
    if not b.coffee_break:
        return "Нет"
    bits: list[str] = []
    if b.coffee_headcount:
        bits.append(f"{b.coffee_headcount} кофе-брейк.")
    if b.coffee_type:
        t = _COFFEE_TYPE_RU.get(b.coffee_type, b.coffee_type)
        if b.coffee_type == "other" and b.coffee_other:
            t = f"другое: {b.coffee_other}"
        bits.append(t)
    if b.foreign_guests:
        bits.append("гости иностранцы")
    return f"Да ({'; '.join(bits)})" if bits else "Да"


def _result_text(b: Booking) -> str:
    label = RU_OUTCOME.get(b.result_outcome or "", "")
    if label and b.result_note:
        return f"{label}: {b.result_note}"
    return label or (b.result_note or "")


def _duration_hours(b: Booking) -> float:
    s, e = _as_utc(b.starts_at), _as_utc(b.ends_at)
    if not s or not e:
        return 0.0
    return round((e - s).total_seconds() / 3600, 2)


def _write_bookings_sheet(ws, bookings: list[Booking]) -> None:
    ws.title = "Бронирования"
    ws.append([h for h, _, _ in _COLUMNS])
    _style_header(ws, len(_COLUMNS))

    for b in bookings:
        s = _as_utc(b.starts_at)
        e = _as_utc(b.ends_at)
        ws.append([
            b.id,
            RU_STATUS.get(b.status, b.status.value),
            "Да" if b.is_urgent else "—",
            b.room.zone.name if b.room and b.room.zone else "—",
            b.room.name if b.room else "—",
            s.strftime("%d.%m.%Y") if s else "—",
            s.strftime("%H:%M") if s else "—",
            e.strftime("%H:%M") if e else "—",
            _duration_hours(b),
            b.event_type,
            b.event_name,
            b.description or "",
            b.attendees,
            _coffee_text(b),
            b.company,
            b.contact_name,
            b.phone,
            b.customer_telegram_id,
            f"@{b.customer_username}" if b.customer_username else "—",
            b.reject_reason or "",
            _result_text(b),
            b.feedback.rating if b.feedback else "",
            (b.feedback.comment or "") if b.feedback else "",
            _fmt_dt(b.created_at),
            _fmt_dt(b.updated_at),
        ])
        row = ws.max_row
        for idx, (_, _, align) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=row, column=idx)
            cell.alignment = align
            cell.border = _BORDER
        if b.is_urgent and b.status != BookingStatus.archived:
            ws.cell(row=row, column=3).fill = _URGENT_FILL

    for idx, (_, width, _) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


_HISTORY_HEADERS = ["№ заявки", "Когда", "Из статуса", "В статус", "Кто (Telegram ID)", "Заметка"]
_HISTORY_WIDTHS = [10, 18, 16, 16, 18, 40]


def _write_history_sheet(ws, bookings: list[Booking]) -> None:
    ws.title = "История статусов"
    ws.append(_HISTORY_HEADERS)
    _style_header(ws, len(_HISTORY_HEADERS))

    rows = [
        (b.id, h.created_at, h.from_status, h.to_status, h.actor_telegram_id, h.note)
        for b in bookings
        for h in b.status_history
    ]
    rows.sort(key=lambda r: _as_utc(r[1]) or datetime.min.replace(tzinfo=timezone.utc))
    for bid, when, frm, to, actor, note in rows:
        ws.append([
            bid,
            _fmt_dt(when),
            RU_STATUS.get(frm, "—") if frm else "—",
            RU_STATUS.get(to, to.value) if to else "—",
            actor if actor is not None else "—",
            note or "",
        ])
        r = ws.max_row
        for col in range(1, len(_HISTORY_HEADERS) + 1):
            ws.cell(row=r, column=col).border = _BORDER
            ws.cell(row=r, column=col).alignment = _WRAP_TOP

    for idx, width in enumerate(_HISTORY_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def period_bounds(date_from: date | None, date_to: date | None) -> list:
    """SQL filters for ``Booking.starts_at`` within an inclusive [date_from, date_to] day range."""
    filters = []
    if date_from is not None:
        filters.append(Booking.starts_at >= datetime.combine(date_from, time.min, tzinfo=timezone.utc))
    if date_to is not None:
        filters.append(
            Booking.starts_at < datetime.combine(date_to, time.min, tzinfo=timezone.utc) + timedelta(days=1)
        )
    return filters


async def build_bookings_workbook(
    session: AsyncSession, date_from: date | None = None, date_to: date | None = None
) -> bytes:
    """Return a styled .xlsx of bookings (optionally limited to a day range) + history."""
    stmt = (
        select(Booking)
        .options(
            selectinload(Booking.room).selectinload(Room.zone),
            selectinload(Booking.status_history),
            selectinload(Booking.feedback),
        )
        .where(*period_bounds(date_from, date_to))
        .order_by(Booking.starts_at.desc())
    )
    bookings = list((await session.execute(stmt)).scalars().all())

    wb = Workbook()
    _write_bookings_sheet(wb.active, bookings)
    _write_history_sheet(wb.create_sheet(), bookings)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def report_filename(now: datetime | None = None) -> str:
    now = now or datetime.now(timezone.utc)
    return f"ats_bookings_{now:%Y-%m-%d}.xlsx"
